from datetime import datetime

from openpyxl import load_workbook


MESES = {
    1: "jan",
    2: "fev",
    3: "mar",
    4: "abr",
    5: "mai",
    6: "jun",
    7: "jul",
    8: "ago",
    9: "set",
    10: "out",
    11: "nov",
    12: "dez",
}


class ExcelImporter:

    CABECALHO = (
        "DATA",
        "DIA",
        "HORA*",
        "PROCEDIMENTOS",
        "EXECUTOR",
        "SISTEMA",
    )

    SISTEMAS_VALIDOS = {
        "SIPPES",
        "SIAPPES",
    }

    def __init__(self, arquivo):

        self.arquivo = arquivo

    @staticmethod
    def _texto(valor):

        if valor is None:
            return ""

        return str(valor).strip()

    @staticmethod
    def _formatar_data(valor):

        if valor is None:
            return ""

        if isinstance(valor, datetime):

            return (
                f"{valor.day:02d}/"
                f"{MESES[valor.month]}/"
                f"{str(valor.year)[2:]}"
            )

        texto = str(valor).strip()

        if "/" in texto:

            partes = texto.split("/")

            if len(partes) == 3:

                mes = partes[1]

                if mes.isdigit():

                    numero_mes = int(mes)

                    if numero_mes in MESES:

                        return (
                            f"{int(partes[0]):02d}/"
                            f"{MESES[numero_mes]}/"
                            f"{partes[2][-2:]}"
                        )

        return texto.lower()

    @staticmethod
    def _formatar_hora(valor):

        if valor is None:
            return ""

        if isinstance(valor, datetime):

            return valor.strftime("%H:%M")

        return str(valor).strip()

    def _detectar_sistema(self, planilha):

        texto_cabecalho = []

        for linha in range(1, 11):

            for coluna in range(1, 10):

                valor = self._texto(
                    planilha.cell(
                        row=linha,
                        column=coluna,
                    ).value
                ).upper()

                if valor:

                    texto_cabecalho.append(valor)

        texto = " ".join(texto_cabecalho)

        if "SIAPPES" in texto:

            return "SIAPPES"

        if "SIPPES" in texto:

            return "SIPPES"

        return ""

    def importar(self):

        workbook = load_workbook(
            filename=self.arquivo,
            data_only=True,
        )

        planilha = workbook.active

        sistema_arquivo = self._detectar_sistema(
            planilha
        )

        cabecalho = tuple(

            self._texto(
                planilha.cell(
                    row=10,
                    column=col,
                ).value
            ).upper()

            for col in range(1, 7)

        )

        cabecalho_sem_sistema = (
            "DATA",
            "DIA",
            "HORA*",
            "PROCEDIMENTOS",
            "EXECUTOR",
            "",
        )

        if (
            cabecalho != self.CABECALHO
            and cabecalho != cabecalho_sem_sistema
        ):

            raise ValueError(
                "Layout do cronograma inválido. "
                "Esperado: DATA, DIA, HORA*, PROCEDIMENTOS, "
                "EXECUTOR, SISTEMA."
            )

        itens = []

        linha = 11

        while True:

            data = planilha.cell(
                linha,
                1,
            ).value

            if data is None:

                break

            descricao = self._texto(
                planilha.cell(
                    linha,
                    4,
                ).value
            )

            if descricao == "":

                break

            sistema_coluna = self._texto(
                planilha.cell(
                    linha,
                    6,
                ).value
            ).upper()

            if sistema_arquivo:

                sistema = sistema_arquivo

            else:

                sistema = sistema_coluna

            if sistema and sistema not in self.SISTEMAS_VALIDOS:

                raise ValueError(
                    f"Sistema inválido na linha {linha}: "
                    f"'{sistema}'. Use SIPPES ou SIAPPES."
                )

            if not sistema:

                raise ValueError(
                    f"Sistema não identificado na linha {linha}. "
                    "Informe SIPPES ou SIAPPES no título do arquivo "
                    "ou na coluna SISTEMA."
                )

            itens.append(

                {

                    "data": self._formatar_data(data),

                    "dia_semana": self._texto(
                        planilha.cell(
                            linha,
                            2,
                        ).value
                    ),

                    "horario": self._formatar_hora(
                        planilha.cell(
                            linha,
                            3,
                        ).value
                    ),

                    "descricao": descricao,

                    "executor": self._texto(
                        planilha.cell(
                            linha,
                            5,
                        ).value
                    ),

                    "sistema": sistema,

                    "cor": None,

                }

            )

            linha += 1

        return itens